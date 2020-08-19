from openpyxl import Workbook

import json
import matplotlib.pyplot as plt    #引入函式庫
import numpy as np
import math


qScoresFile = "D:\\g1080265\\instagram\\questionnaire_scores.json"
with open(qScoresFile, 'r') as load_f:
    qScores = json.load(load_f)

animalDataFile = "D:\\g1080265\\instagram\\animal.json"
with open(animalDataFile, 'r') as load_f:
    animalScore100 = json.load(load_f)

vehicleDataFile = "D:\\g1080265\\instagram\\vehicle.json"
with open(vehicleDataFile, 'r') as load_f:
    vehicleScore100 = json.load(load_f)

WV = Workbook()
WN = Workbook()

WVResult = WV.active
WNResult = WN.active

WVResult.append(['編號','帳號','問卷數量','問卷dog',  '問卷cat', 'Word2Vecdog', 'Word2Veccat','距離'])
WNResult.append(['編號','帳號','問卷數量','問卷dog',  '問卷cat', 'Wordnetdog', 'Wordnetcat','距離'])

animalUsers = [ line.rstrip('\n') for line in open('D:\\g1080265\\instagram\\questionnaire_animal.txt', 'r') if line.rstrip('\n') != '']

color = ['m', 'y', 'g', 'r', 'c', 'm', 'y', 'k', 'w']
marker = []

i = 0
for user in animalUsers:
    # print( qScores[user] )
    i += 1

    qDog = round( (sum( qScores[user]['dog'] )/len(qScores[user]['dog'])),2 )
    qCat = round( (sum( qScores[user]['cat'] )/len(qScores[user]['cat'])),2 )
    wvDog = animalScore100[user]['word2vec']["dog"]
    wvCat = animalScore100[user]['word2vec']["cat"]
    wnDog = animalScore100[user]['wordnet']["dog"]
    wnCat = animalScore100[user]['wordnet']["cat"]

    VD =  round( (math.sqrt((qDog-wvDog)*(qDog-wvDog) + (qCat-wvCat)*(qCat-wvCat))), 2)
    ND =  round( (math.sqrt((qDog-wnDog)*(qDog-wnDog) + (qCat-wnCat)*(qCat-wnCat))), 2)

    print([i, user, len(qScores[user]['dog']), len(qScores[user]['cat']), qDog, qCat, wvDog, wvCat, VD])
    WVResult.append([i, user, len(qScores[user]['dog']), qDog, qCat, wvDog, wvCat, VD])

    # print([i, user, len(qScores[user]['dog']), len(qScores[user]['cat']), qDog, qCat, wnDog, wnCat, ND])
    WNResult.append([i, user, len(qScores[user]['dog']), qDog, qCat, wnDog, wnCat, ND])

    plt.xlim(0, 100)
    plt.ylim(0, 100)
    
    #设置坐标轴刻度
    my_x_ticks = np.arange(0, 101, 10)
    my_y_ticks = np.arange(0, 101, 10)
    plt.xticks(my_x_ticks)
    plt.yticks(my_y_ticks)

    # 
    plt.scatter(qDog, qCat, s=100, c=color[i%5], marker='o')
    # plt.annotate("({},{})".format(qDog, qCat), xy=(qDog, qCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point
    plt.annotate(i, xy=(qDog, qCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.scatter(wvDog, wvCat, s=100, c=color[i%5], marker='*')
    # plt.annotate("({},{})".format(wvDog, wvCat), xy=(wvDog, wvCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point
    plt.annotate(i, xy=(wvDog, wvCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.scatter(wnDog, wnCat, s=100, c=color[i%5], marker='^')
    # plt.annotate("({},{})".format(wnDog, wnCat), xy=(wnDog, wnCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point
    plt.annotate(i, xy=(wnDog, wnCat), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.xlabel("dog")               #設定X軸名稱
    plt.ylabel("cat")               #設定Y軸名稱



    if i == 10:
        
        plt.title("Animal Questionnaire")
        
        plt.scatter(120, 120, label='Questionare', c='b', s=100, marker='o')
        plt.scatter(120, 120, label='Word2Vec', c='b', s=100, marker='*')
        plt.scatter(120, 120, label='Wordnet', c='b', s=100, marker='^')
        # legend position 待處理
        plt.legend(loc='lower left')
        plt.show()                    #呈現所繪圖表

WV.save("WVAnimalResult.xlsx")
WN.save("WNAnimalResult.xlsx")





vehicleDataFile = "D:\\g1080265\\instagram\\vehicle.json"
with open(vehicleDataFile, 'r') as load_f:
    vehicleScore100 = json.load(load_f)

WV = Workbook()
WN = Workbook()

WVResult = WV.active
WNResult = WN.active

WVResult.append(['編號','帳號','問卷數量','問卷car',  '問卷motorcycle', 'Word2Veccar', 'Word2Vecmotorcycle','距離'])
WNResult.append(['編號','帳號','問卷數量','問卷car',  '問卷motorcycle', 'Wordnetcar', 'Wordnetmotorcycle','距離'])

vehicleUsers = [ line.rstrip('\n') for line in open('D:\\g1080265\\instagram\\questionnaire_vehicle.txt', 'r') if line.rstrip('\n') != '']

i = 0
for user in vehicleUsers:
    # print( qScores[user] )
    i += 1

    qCar = round( (sum( qScores[user]['car'] )/len(qScores[user]['car'])),2 )
    qMotor = round( (sum( qScores[user]['motorcycle'] )/len(qScores[user]['motorcycle'])),2 )
    wvCar = vehicleScore100[user]['word2vec']["car"]
    wvMotor = vehicleScore100[user]['word2vec']["motorcycle"]
    wnCar = vehicleScore100[user]['wordnet']["car"]
    wnMotor = vehicleScore100[user]['wordnet']["motorcycle"]

    VD =  round( (math.sqrt((qCar-wvCar)*(qCar-wvCar) + (qMotor-wvMotor)*(qMotor-wvMotor))), 2)
    ND =  round( (math.sqrt((qCar-wnCar)*(qCar-wnCar) + (qMotor-wnMotor)*(qMotor-wnMotor))), 2)

    print([i, user, len(qScores[user]['car']), len(qScores[user]['motorcycle']), qCar, qMotor, wvCar, wvCat, VD])
    WVResult.append([i, user, len(qScores[user]['car']), qCar, qMotor, wvCar, wvMotor, VD])

    # print([i, user, len(qScores[user]['car']), len(qScores[user]['motorcycle']), qDog, qMotor, wnDog, wnCat, ND])
    WNResult.append([i, user, len(qScores[user]['car']), qCar, qMotor, wnCar, wnMotor, ND])

    plt.xlim(0, 100)
    plt.ylim(0, 100)
    
    #设置坐标轴刻度
    my_x_ticks = np.arange(0, 101, 10)
    my_y_ticks = np.arange(0, 101, 10)
    plt.xticks(my_x_ticks)
    plt.yticks(my_y_ticks)

    # 
    plt.scatter(qCar, qMotor, s=100, c=color[i%5], marker='o')
    plt.annotate(i, xy=(qCar, qMotor), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.scatter(wvCar, wvMotor, s=100, c=color[i%5], marker='*')
    plt.annotate(i, xy=(wvCar, wvMotor), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.scatter(wnCar, wnMotor, s=100, c=color[i%5], marker='^')
    plt.annotate(i, xy=(wnCar, wnMotor), xytext=(5, 0), textcoords='offset points',ha='left', va='center') #label on each point

    plt.xlabel("car")               #設定X軸名稱
    plt.ylabel("motorcycle")               #設定Y軸名稱


    if i == 10:
        
        plt.title("Vehicle Questionnaire")
        
        plt.scatter(120, 120, label='Questionare', c='b', s=100, marker='o')
        plt.scatter(120, 120, label='Word2Vec', c='b', s=100, marker='*')
        plt.scatter(120, 120, label='Wordnet', c='b', s=100, marker='^')
        # legend position 待處理
        plt.legend(loc='lower left')
        plt.show()                    #呈現所繪圖表

WV.save("WVVehicleResult.xlsx")
WN.save("WNVehicleResult.xlsx")


# from matplotlib import pylab
# import numpy as np
# import random

# batch = np.ndarray(shape=(10,2), dtype=np.int32)
# for i in range(10):
#     batch[i,0]= random.randint(0, 20)
#     batch[i,1]= random.randint(0, 20)
# string = "1 2 3 4 5 6 7 8 9 10"
# name = string.split()

# pylab.figure(figsize=(10,5)) #figsize=(width,height) by inch
# for i, label in enumerate(name):
#     x,y=batch[i,:]
#     pylab.scatter(x,y) #plot scatter point
#     pylab.annotate(label, xy=(x, y), xytext=(5, 2), textcoords='offset points',ha='right', va='bottom') #label on each point
# pylab.show()

# pylab.figure(figsize=(10,5)) : 決定圖的大小
# pylab.scatter(x,y) : 將點以x,y座標畫在圖上
# pylab.annotate : 將文字映在圖上
# label ：為映的文字
# xy ：文字的座標
# xytext : 文字對xy座標的平移量，數字愈大移的愈遠
# va : 有'center'   'left'   'right'  三種選擇，意思是選擇用文字方塊的 中間/左邊/右邊對齊xy座標點
# ha: 有 'center'   'top'   'bottom' 'baseline，意思也是用文字方塊的上/中/下緣對齊座標，
# baseline 目前不知用處為何